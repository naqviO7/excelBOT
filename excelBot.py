import time
import numpy as np
from openpyxl import workbook
import pandas as pd
import openpyxl as xl
from os import system
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


#function to display banner 
def Banner_of_ExcelBOT():
    print("""
            ____     _ _ ____   ___ _____
  _____  __/ ___|___| | | __ ) / _ \_   _|
 / _ \ \/ / |   / _ \ | |  _ \| | | || |
|  __/>  <| |__|  __/ | | |_) | |_| || |
 \___/_/\_\\____\___|_|_|____/ \___/ |_|

    """)


#function containing main menun of excel bot
def Main_Menu_of_ExcelBOT():
    print(' ------------------------------------ ')
    print('|          M A I N M E N U           |')
    print(' ------------------------------------ ')
    print('| [=] Options to Perform Operations  |')
    print('| --> 1 Manually Perform Excel Tasks |')
    print('| --> 2 Automate Excel Tasks         |')
    print('| --> 0 Quit                         |')
    print(' ------------------------------------ ')
    print('\n')


#class to perform excell tasks manuall
class Maunal_ExcelBOT_Tasks:
    #menu to show options for manuall tasks
    def Manual_Funcs_Menu(self):
        print(' ------------------------------------- ')
        print('|          MANUALL MENU               |')
        print(' ------------------------------------- ')
        print('| [+] Options to Perform Operations   |')
        print('| --> 1 Create New Excel File         |')
        print('| --> 2 Merge Cells of Excel File     |') 
        print('| --> 3 Copy and Pase Excel File Data |')
        print('| --> 4 Display Excel File Data       |')
        print('| --> 0 Quit                          |')
        print(' ------------------------------------- ')
        print('\n')
    

    #function to create new excel file
    def Create_ExcelFile(self):
        print('-'*15,'Create Excel File','-'*15)
    
        excel_filename=input('Enter File Name to Create: ')
        excel_file=Workbook()
        excel_file.save(excel_filename)
    
        print(f'{excel_filename} is successfully created!')
    
    
    #function to display excel file data
    def Display_ExcelFile_Data(self):
        print('-'*15,'Display Excel File Data','-'*15)
    
        excel_file_name=input('Enter File Name: ')
        excel_sheet_name=input('Enter Excel Sheet Name: ')
        df_excel_file=pd.read_excel(excel_file_name,sheet_name=excel_sheet_name)
    
        print(df_excel_file)
    
    
    #function to merge cells
    def Merge_Excel_Cells(self):
        print('-'*15,'Merge Excel Cells','-'*15)
        excel_file_name=input('Enter File Name: ')
    
        work_book=load_workbook(excel_file_name)
        work_sheet=work_book.active

        #range shoudl be in starting_cells:ending_cells
        range=input('Enter Range of Cells to Merge: ')
    
        work_sheet.merge_cells(range)
        work_book.save(excel_file_name)
    
        print(f'{range} Cells are Merged!')
    
    
    #function to copy and paste data
    def Copy_Paste_Excel_Data(self):
        print('-'*15,'Copy Paste Excel Data','-'*15)
        source_file=input('Enter Name of Source File: ')
        destination_file=input('Enter Name of Destination File: ')
        work_book=Workbook()
        work_book.save(destination_file)
        
        src_wb = load_workbook(source_file)
        dest_wb = load_workbook(destination_file)

        src_sheet_name=input('Enter Source Sheet Name: ')
        dest_sheet_name=input('Enter Destination Sheet Name: ')
        
        src_sheet = src_wb.get_sheet_by_name(src_sheet_name)
        dest_sheet = dest_wb.get_sheet_by_name(dest_sheet_name)

        for i in range(1, src_sheet.max_row+1):
            for j in range(1, src_sheet.max_column+1):
                dest_sheet.cell(row=i, column=j).value = src_sheet.cell(row=i, column=j).value

        src_wb.save(source_file)
        dest_wb.save(destination_file)
        
        print(f'Data is Successfully Copied to {destination_file}!')



#class for automating excel tasks
class Automated_ExcelBOT_Tasks:
    #function to display B,D,G,H column data
    def Display_Column_Data(self):
        print('-'*15,'Displaying Column Data','-'*15)

        work_book=load_workbook('colsfile.xlsx')
        work_sheet=work_book.active
        
        colb=work_sheet['B']
        cold=work_sheet['D']
        colg=work_sheet['G']
        colh=work_sheet['H']
        
        print('Data in Column B')
        for cell in colb:
            print(cell.value)
        print('\n')
            
        print('Data in Column D')
        for cell in cold:
            print(cell.value)
        print('\n')
            
        print('Data in Column G')
        for cell in colg:
            print(cell.value)   
        print('\n')
        
        print('Data in Column H')
        for cell in colh:
            print(cell.value)         
        print('\n')
    
        print('Combining G and H Columns!')
        work_sheet.merge_cells('G3:H3')
        work_sheet.merge_cells('G4:H4')
        print('G and H Columns are Combined and Saved in a New File!')
        work_book.save('merged_cols.xlsx')
        print('\n')
    
    
    #functon to extract value from excel file
    def  Seperate_Excel_File(self):
        print('-'*15,'Seperating Excel File','-'*15)
        exl_df=pd.read_excel('convert.xlsx')
        print(exl_df)
        conv=exl_df['Name'].where(exl_df['Mon 28/06/2021']=='15:00-23:00')
        conv=exl_df['Name'].where(exl_df['Tue 29/06/2021']=='9:00-17:00')  
        print(conv.dropna())       
        
        wb=load_workbook('convert.xlsx')
        ws=wb.active
        print(ws['D1'].value,'\t\t\t',ws['F1'].value,ws['G1'].value)  
        print(ws['D12'].value,ws['F12'].value,'\t',ws['G12'].value) 
        wb.save('new_convert.xlsx') 
       
    
    #function to display simple file data    
    def Simple_File_data(self):
        print('-'*15,'Simple Excel File Data','-'*15)
        excel_df=pd.read_excel('simplfile.xlsx')
        print('Printing File Data\n')
        print(excel_df.dropna())
    
    
    #function to extract monday sheet data
    def Complex_Excel_File(self):
        pass 

#managing for import system      
if __name__=='__main__':
    system('cls')
    time.sleep(3)
    Banner_of_ExcelBOT()
    
    time.sleep(2)
    Main_Menu_of_ExcelBOT()
    
    time.sleep(1) 
    
    key=int(input('Enter Number to Perform Operation: '))
    if key==1:
        manual_funcs=Maunal_ExcelBOT_Tasks()
        time.sleep(2)
        manual_funcs.Manual_Funcs_Menu()
        func_key=int(input('Enter Number to Perform Operation: '))
        
        if func_key==1:
            time.sleep(2)
            manual_funcs.Create_ExcelFile()
        elif func_key==2:
            time.sleep(2)
            manual_funcs.Merge_Excel_Cells()
        elif func_key==3:
            time.sleep(2)
            manual_funcs.Copy_Paste_Excel_Data()
        elif func_key==4:
            time.sleep(2)
            manual_funcs.Display_ExcelFile_Data()
        elif func_key==0:
            time.sleep(2)
            system('cls')
            exit(0)
    
    elif key==2:
        time.sleep(1)
        automated_funcs=Automated_ExcelBOT_Tasks()
        time.sleep(2)
        automated_funcs.Seperate_Excel_File()
        time.sleep(2)
        automated_funcs.Display_Column_Data()
        time.sleep(2)
        automated_funcs.Simple_File_data()
        
    elif key==0:
        print('-'*15,'Quiting excelBOT','-'*15)
        time.sleep(1)
        system('cls')
        exit() 
#ENDOFCODE
